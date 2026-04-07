from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
import copy

# ── Palette ─────────────────────────────────────────────────────────────────
VLAN_MAP = {
    "BROADCAST":        {"color": "185FA5", "desc": "Broadcast / Difusión",    "short": "BCAST"},
    "GMS":              {"color": "639922", "desc": "Gestión General",          "short": "GMS"  },
    "GMS2":             {"color": "BA7517", "desc": "Gestión Secundaria",       "short": "GMS2" },
    "ACCESS POINT":     {"color": "0F6E56", "desc": "Red Inalámbrica (AP)",     "short": "AP"   },
    "SERVER MONITOREO": {"color": "A32D2D", "desc": "Servidor de Monitoreo",    "short": "SRV"  },
    "Sin asignar":      {"color": "888780", "desc": "Sin asignar / Vacío",      "short": "—"    },
    "Desasignar":       {"color": "5F5E5A", "desc": "Desasignado / Para retirar","short":"DEL"  },
}

SWITCHES = [
    {
        "id": "FSW_124F_Dist",
        "nombre": "FortiSwitch 124F",
        "modelo": "FortiSwitch 124F",
        "ubicacion": "Venue Principal (IBC) — Distribución",
        "ip": "10.0.1.100",
        "puertos": [
            ("P01","PT13","BROADCAST","activo"),
            ("P02","PT17","GMS","activo"),
            ("P03","PT09","BROADCAST","activo"),
            ("P04","PT07","GMS","activo"),
            ("P05","PT08","GMS2","activo"),
            ("P06","PT19","GMS2","activo"),
            ("P07","PT21","BROADCAST","activo"),
            ("P08","","Sin asignar","vacio"),
            ("P09","PT03","Desasignar","desasignado"),
            ("P10","","Sin asignar","vacio"),
            ("P11","PT01","Desasignar","desasignado"),
            ("P12","PT02","Desasignar","desasignado"),
            ("P13","PT11","GMS2","activo"),
            ("P14","PT18","BROADCAST","activo"),
            ("P15","PT20","BROADCAST","activo"),
            ("P16","PT25","GMS2","activo"),
            ("P17","PT15","GMS","activo"),
            ("P18","PT16","GMS2","activo"),
            ("P19","PT32","GMS","activo"),
            ("P20","PT27","GMS2","activo"),
            ("P21","PT28","GMS2","activo"),
            ("P22","PT31","GMS2","activo"),
            ("P23","PT24","GMS","activo"),
            ("P24","PT29","GMS","activo"),
        ]
    },
    {
        "id": "FSW_624F_IBC_A",
        "nombre": "FortiSwitch 624F IBC A",
        "modelo": "FortiSwitch 624F",
        "ubicacion": "IBC — Centro de Control A",
        "ip": "10.0.2.100",
        "puertos": [
            ("P01","PT14","GMS2","activo"),
            ("P02","PT04","BROADCAST","activo"),
            ("P03","PT12","GMS","activo"),
            ("P04","PT06","GMS2","activo"),
            ("P05","PT22","GMS","activo"),
            ("P06","PT05","GMS","activo"),
            ("P07","PT26","GMS","activo"),
            ("P08","PT30","GMS","activo"),
            ("P09","PT33","GMS2","activo"),
            ("P10","Reservado","SERVER MONITOREO","reservado"),
            ("P11","PT10","BROADCAST","activo"),
            ("P12","KI01","GMS2","activo"),
            ("P13","PT23","GMS2","activo"),
            ("P14","KI02","GMS","activo"),
            ("P15","Reservado","BROADCAST","reservado"),
            ("P16","KI03","BROADCAST","activo"),
            ("P17","KI04","BROADCAST","activo"),
            ("P18","KI05","GMS","activo"),
            ("P19","KI06","GMS2","activo"),
            ("P20","KI07","GMS","activo"),
            ("P21","","Sin asignar","vacio"),
            ("P22","","Sin asignar","vacio"),
            ("P23","","Sin asignar","vacio"),
            ("P24","","Sin asignar","vacio"),
        ]
    },
    {
        "id": "FSW_624F_IBC_B",
        "nombre": "FortiSwitch 624F IBC B",
        "modelo": "FortiSwitch 624F",
        "ubicacion": "IBC — Centro de Control B",
        "ip": "10.0.3.100",
        "puertos": [
            ("P01","Reservado","ACCESS POINT","reservado"),
            ("P02","Reservado","ACCESS POINT","reservado"),
            ("P03","KI08","GMS2","activo"),
            ("P04","KI09","GMS","activo"),
            ("P05","KI10","GMS","activo"),
            ("P06","KI11","GMS2","activo"),
            ("P07","KI12","BROADCAST","activo"),
            ("P08","KI13","BROADCAST","activo"),
            ("P09","KI14","GMS","activo"),
            ("P10","KI15","GMS2","activo"),
            ("P11","KI16","BROADCAST","activo"),
            ("P12","KI17","BROADCAST","activo"),
            ("P13","","Sin asignar","vacio"),
            ("P14","","Sin asignar","vacio"),
            ("P15","","Sin asignar","vacio"),
            ("P16","","Sin asignar","vacio"),
            ("P17","","Sin asignar","vacio"),
            ("P18","","Sin asignar","vacio"),
            ("P19","","Sin asignar","vacio"),
            ("P20","","Sin asignar","vacio"),
            ("P21","","Sin asignar","vacio"),
            ("P22","","Sin asignar","vacio"),
            ("P23","","Sin asignar","vacio"),
            ("P24","","Sin asignar","vacio"),
        ]
    },
]

# ── Style helpers ────────────────────────────────────────────────────────────
def ft(bold=False, size=11, color="000000", name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="D0D3DC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row, cols, bg="1A1F2E", fg="FFFFFF", height=22):
    ws.row_dimensions[row].height = height
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = ft(bold=True, size=10, color=fg)
        cell.alignment = center()
        cell.border = border_thin()

def style_data_row(ws, row, col_count, light=True):
    bg = "F7F8FA" if light else "FFFFFF"
    ws.row_dimensions[row].height = 18
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = ft(size=10)
        cell.alignment = left_mid()
        cell.border = border_thin()


# ── Workbook ─────────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 1: INSTRUCCIONES
# ═══════════════════════════════════════════════════════════════════════════
ws_inst = wb.active
ws_inst.title = "Instrucciones"
ws_inst.sheet_view.showGridLines = False
ws_inst.column_dimensions["A"].width = 2
ws_inst.column_dimensions["B"].width = 28
ws_inst.column_dimensions["C"].width = 60

# Header band
for r in range(1, 6):
    ws_inst.row_dimensions[r].height = 14
    for c in range(1, 12):
        ws_inst.cell(row=r, column=c).fill = fill("1A1F2E")

ws_inst.merge_cells("B2:J4")
t = ws_inst["B2"]
t.value = "JSJ 2026 — Visualizador de Switches · Guía de Uso"
t.font = ft(bold=True, size=18, color="FFFFFF")
t.alignment = Alignment(horizontal="left", vertical="center")

ws_inst.merge_cells("B5:J5")
sub = ws_inst["B5"]
sub.value = "FortiSwitch Network Management · IBC Venues · v1.0 · 2026-04-07"
sub.font = ft(size=10, color="8899BB")
sub.alignment = Alignment(horizontal="left", vertical="center")

# Spacer
ws_inst.row_dimensions[6].height = 12

sections = [
    ("📋 ESTRUCTURA DEL ARCHIVO", "1A1F2E", "FFFFFF", [
        ("Hoja Resumen",        "Vista consolidada de todos los switches, estadísticas por VLAN y estado."),
        ("FSW_124F_Dist",       "Datos de puertos del FortiSwitch 124F (Distribución) — 24 puertos."),
        ("FSW_624F_IBC_A",      "Datos de puertos del FortiSwitch 624F IBC A — 24 puertos."),
        ("FSW_624F_IBC_B",      "Datos de puertos del FortiSwitch 624F IBC B — 24 puertos."),
        ("VLANs",               "Catálogo de VLANs con colores, descripciones y conteos automáticos."),
    ]),
    ("✏️ CÓMO EDITAR DATOS", "185FA5", "FFFFFF", [
        ("Columna Puerto",      "No modificar. Identificador fijo: P01–P24."),
        ("Columna Cable",       "Escribe el código de cable (ej: PT13, KI08) o deja vacío."),
        ("Columna VLAN",        "Usa el desplegable de validación para elegir una VLAN válida."),
        ("Columna Estado",      "Usa el desplegable: activo · vacio · desasignado · reservado."),
        ("Columna Dispositivo", "Opcional. Nombre o descripción del equipo conectado."),
        ("Columna Notas",       "Campo libre para observaciones adicionales del técnico."),
    ]),
    ("☁️ FLUJO ONEDRIVE → VISUALIZADOR", "0F6E56", "FFFFFF", [
        ("1. Editar en OneDrive", "Abre este archivo desde OneDrive en Excel Web o Excel Desktop."),
        ("2. Actualizar datos",   "Modifica cables, VLANs y estados según la instalación real."),
        ("3. Descargar / Guardar","Guarda como .xlsx (sin cambiar nombre ni estructura de hojas)."),
        ("4. Importar al visualizador", "En la app web (jsj2026-switches.html) usa el botón '📂 Importar Excel'."),
        ("5. Verificar cambios",  "El visualizador leerá automáticamente las hojas de switches."),
    ]),
    ("⚠️ REGLAS IMPORTANTES", "A32D2D", "FFFFFF", [
        ("No renombrar hojas",   "Los nombres FSW_124F_Dist, FSW_624F_IBC_A, FSW_624F_IBC_B son fijos."),
        ("No mover columnas",    "El orden A–H debe mantenerse para que el parser funcione correctamente."),
        ("No borrar fila 1",     "La fila de encabezados es requerida por el sistema de importación."),
        ("VLANs válidas",        "Usar solo las VLANs del catálogo (hoja VLANs). Respetar mayúsculas."),
        ("Guardar como .xlsx",   "No guardar como .xls ni .csv. El visualizador requiere formato .xlsx."),
    ]),
]

row = 7
for title, hdr_color, hdr_fg, items in sections:
    ws_inst.row_dimensions[row].height = 20
    ws_inst.merge_cells(f"B{row}:J{row}")
    c = ws_inst[f"B{row}"]
    c.value = title
    c.font = ft(bold=True, size=11, color=hdr_fg)
    c.fill = fill(hdr_color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    row += 1

    for key, val in items:
        ws_inst.row_dimensions[row].height = 28
        kc = ws_inst.cell(row=row, column=2)
        kc.value = key
        kc.font = ft(bold=True, size=10, color="1A1F2E")
        kc.fill = fill("EEF1F8")
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        kc.border = border_thin()

        vc = ws_inst.cell(row=row, column=3)
        vc.value = val
        vc.font = ft(size=10)
        vc.fill = fill("F7F8FA")
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        vc.border = border_thin()
        # merge C across to J
        ws_inst.merge_cells(f"C{row}:J{row}")
        row += 1

    row += 1  # spacer


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET: VLANS
# ═══════════════════════════════════════════════════════════════════════════
ws_vlan = wb.create_sheet("VLANs")
ws_vlan.sheet_view.showGridLines = False

col_w = [2, 22, 36, 14, 14, 14, 14, 14]
col_keys = ["", "A", "B", "C", "D", "E", "F", "G", "H"]
for i, w in enumerate(col_w, 1):
    ws_vlan.column_dimensions[get_column_letter(i)].width = w

# Banner
for r in range(1, 5):
    ws_vlan.row_dimensions[r].height = 13
    for c in range(1, 10):
        ws_vlan.cell(row=r, column=c).fill = fill("1A1F2E")

ws_vlan.merge_cells("B2:H3")
t = ws_vlan["B2"]
t.value = "Catálogo de VLANs — JSJ 2026"
t.font = ft(bold=True, size=16, color="FFFFFF")
t.alignment = Alignment(horizontal="left", vertical="center")

row = 5
ws_vlan.row_dimensions[row].height = 12

# Headers
row = 6
headers = ["VLAN", "Descripción", "Color Hex", "Color Visual", "Prioridad", "Switch A", "Switch B", "Switch C"]
ws_vlan.row_dimensions[row].height = 22
for ci, h in enumerate(headers, 2):
    c = ws_vlan.cell(row=row, column=ci)
    c.value = h
    c.font = ft(bold=True, size=10, color="FFFFFF")
    c.fill = fill("1A1F2E")
    c.alignment = center()
    c.border = border_thin()
row += 1

PRIORITY = {
    "BROADCAST": "Normal", "GMS": "Alta", "GMS2": "Alta",
    "ACCESS POINT": "Normal", "SERVER MONITOREO": "Crítica",
    "Sin asignar": "—", "Desasignar": "—"
}

# Pre-compute counts per switch for each vlan
def count_vlan(switch, vlan_name):
    return sum(1 for p in switch["puertos"] if p[2] == vlan_name)

for vi, (vname, vinfo) in enumerate(VLAN_MAP.items()):
    ws_vlan.row_dimensions[row].height = 20
    bg_row = "F7F8FA" if vi % 2 == 0 else "FFFFFF"

    cells = [vname, vinfo["desc"], f"#{vinfo['color']}", "", PRIORITY.get(vname,"—")]
    for s in SWITCHES:
        cells.append(count_vlan(s, vname))

    for ci, val in enumerate(cells, 2):
        c = ws_vlan.cell(row=row, column=ci)
        c.value = val
        c.border = border_thin()
        c.alignment = center() if ci >= 5 else left_mid()
        if ci == 2:
            c.font = ft(bold=True, size=10, color=vinfo["color"])
            c.fill = fill(bg_row)
        elif ci == 4:
            # color hex cell - dark text
            c.font = ft(size=10, color="555555")
            c.fill = fill(bg_row)
        elif ci == 5:
            # Color visual swatch
            c.fill = fill(vinfo["color"])
            c.value = vinfo["short"]
            c.font = ft(bold=True, size=10, color="FFFFFF")
        else:
            c.font = ft(size=10)
            c.fill = fill(bg_row)

    row += 1


# ═══════════════════════════════════════════════════════════════════════════
#  HELPER: Build a switch sheet
# ═══════════════════════════════════════════════════════════════════════════
VLAN_NAMES = list(VLAN_MAP.keys())
ESTADO_NAMES = ["activo", "vacio", "desasignado", "reservado"]

def build_switch_sheet(wb, sw):
    ws = wb.create_sheet(sw["id"])
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [2, 9, 16, 22, 20, 16, 32, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Banner ──
    for r in range(1, 6):
        ws.row_dimensions[r].height = 13
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = fill("1A1F2E")

    ws.merge_cells("B2:I3")
    t = ws["B2"]
    t.value = f"{sw['nombre']}  ·  {sw['ubicacion']}"
    t.font = ft(bold=True, size=14, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B4:I4")
    sub = ws["B4"]
    sub.value = f"IP: {sw['ip']}  ·  Modelo: {sw['modelo']}  ·  Total Puertos: {len(sw['puertos'])}"
    sub.font = ft(size=10, color="8899BB")
    sub.alignment = Alignment(horizontal="left", vertical="center")

    row = 6

    # ── Meta info box ──
    meta_labels = [
        ("Switch ID", sw["id"]),
        ("Nombre",    sw["nombre"]),
        ("Modelo",    sw["modelo"]),
        ("Ubicación", sw["ubicacion"]),
        ("IP Gestión",sw["ip"]),
        ("SSH Port",  "22"),
        ("Total Puertos", str(len(sw["puertos"]))),
    ]
    ws.row_dimensions[row].height = 12
    row += 1
    meta_start = row
    for label, val in meta_labels:
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row=row, column=2)
        lc.value = label
        lc.font = ft(bold=True, size=9, color="555555")
        lc.fill = fill("EEF1F8")
        lc.border = border_thin()
        lc.alignment = left_mid()

        vc = ws.cell(row=row, column=3)
        vc.value = val
        vc.font = ft(size=10, color="1A1F2E")
        vc.fill = fill("F7F8FA")
        vc.border = border_thin()
        vc.alignment = left_mid()
        ws.merge_cells(f"C{row}:E{row}")
        row += 1

    # Stats on the right side of meta
    stat_col = 7
    stat_row = meta_start

    ws.cell(row=stat_row-1, column=stat_col).value = "ESTADÍSTICAS"
    ws.cell(row=stat_row-1, column=stat_col).font = ft(bold=True, size=9, color="FFFFFF")
    ws.cell(row=stat_row-1, column=stat_col).fill = fill("1A1F2E")
    ws.cell(row=stat_row-1, column=stat_col).alignment = center()
    ws.merge_cells(start_row=stat_row-1, start_column=stat_col, end_row=stat_row-1, end_column=stat_col+1)

    data_start_row = row + 2 + 1  # approx (headers at row+2, data starts row+3)
    # We'll write formulas pointing to data
    # First: figure out where data starts
    # Let's just hardcode the row after we know it
    # We'll build a simple stats block after data section

    row += 1  # spacer

    # ── Headers ──
    header_row = row
    ws.row_dimensions[header_row].height = 22
    col_headers = ["Puerto", "Cable", "VLAN", "Estado", "Dispositivo", "Notas"]
    hdr_colors  = ["2D3250", "2D3250", "2D3250", "2D3250", "2D3250", "2D3250"]
    for ci, (h, hc) in enumerate(zip(col_headers, hdr_colors), 2):
        c = ws.cell(row=header_row, column=ci)
        c.value = h
        c.font = ft(bold=True, size=10, color="FFFFFF")
        c.fill = fill(hc)
        c.alignment = center()
        c.border = border_thin()
    row += 1

    data_start = row
    # ── Data rows ──
    vlan_dv = DataValidation(
        type="list",
        formula1='"' + ','.join(VLAN_NAMES) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    estado_dv = DataValidation(
        type="list",
        formula1='"' + ','.join(ESTADO_NAMES) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(vlan_dv)
    ws.add_data_validation(estado_dv)

    for ri, (port, cable, vlan, estado) in enumerate(sw["puertos"]):
        is_even = ri % 2 == 0
        base_bg = "F7F8FA" if is_even else "FFFFFF"
        ws.row_dimensions[row].height = 20

        # Port number (locked look)
        pc = ws.cell(row=row, column=2, value=port)
        pc.font = ft(bold=True, size=10, color="1A1F2E")
        pc.fill = fill("E8EAF0")
        pc.alignment = center()
        pc.border = border_thin()

        # Cable
        cc = ws.cell(row=row, column=3, value=cable)
        cc.font = ft(size=10, color="1A1F2E" if cable else "AAAAAA", name="Courier New" if cable else "Arial")
        cc.fill = fill(base_bg)
        cc.alignment = center()
        cc.border = border_thin()

        # VLAN — colored
        vinfo = VLAN_MAP.get(vlan, {"color":"888888","desc":"","short":"?"})
        vc = ws.cell(row=row, column=4, value=vlan)
        vc.font = ft(bold=True, size=10, color="FFFFFF")
        vc.fill = fill(vinfo["color"])
        vc.alignment = center()
        vc.border = border_thin()
        vlan_dv.add(vc)

        # Estado
        ec = ws.cell(row=row, column=5, value=estado)
        estado_colors = {
            "activo": ("166534","DCFCE7"), "vacio": ("6B7280","F3F4F6"),
            "desasignado": ("475569","F1F5F9"), "reservado": ("92400E","FEF3C7")
        }
        efg, ebg = estado_colors.get(estado, ("000000", "FFFFFF"))
        ec.font = ft(size=10, color=efg)
        ec.fill = fill(ebg)
        ec.alignment = center()
        ec.border = border_thin()
        estado_dv.add(ec)

        # Dispositivo
        dc = ws.cell(row=row, column=6, value="")
        dc.font = ft(size=10, color="555555")
        dc.fill = fill(base_bg)
        dc.alignment = left_mid()
        dc.border = border_thin()

        # Notas
        nc = ws.cell(row=row, column=7, value="")
        nc.font = ft(size=10, color="555555")
        nc.fill = fill(base_bg)
        nc.alignment = left_mid()
        nc.border = border_thin()

        row += 1

    data_end = row - 1

    # ── STATS block (after data) ──
    row += 1
    ws.row_dimensions[row].height = 12
    row += 1

    stat_items = [
        ("Total Puertos",     f"=COUNTA(B{data_start}:B{data_end})"),
        ("Activos",           f'=COUNTIF(E{data_start}:E{data_end},"activo")'),
        ("Vacíos",            f'=COUNTIF(E{data_start}:E{data_end},"vacio")'),
        ("Desasignados",      f'=COUNTIF(E{data_start}:E{data_end},"desasignado")'),
        ("Reservados",        f'=COUNTIF(E{data_start}:E{data_end},"reservado")'),
        ("VLANs distintas",   f"=SUMPRODUCT(1/COUNTIF(D{data_start}:D{data_end},D{data_start}:D{data_end}))"),
    ]

    ws.merge_cells(f"B{row}:G{row}")
    sh = ws.cell(row=row, column=2)
    sh.value = "RESUMEN AUTOMÁTICO (fórmulas)"
    sh.font = ft(bold=True, size=10, color="FFFFFF")
    sh.fill = fill("185FA5")
    sh.alignment = center()
    row += 1

    for label, formula in stat_items:
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row=row, column=2)
        lc.value = label
        lc.font = ft(bold=True, size=10, color="1A1F2E")
        lc.fill = fill("EEF1F8")
        lc.border = border_thin()
        lc.alignment = left_mid()

        fc = ws.cell(row=row, column=3)
        fc.value = formula
        fc.font = ft(size=10, color="000000")
        fc.fill = fill("F7F8FA")
        fc.border = border_thin()
        fc.alignment = center()
        ws.merge_cells(f"C{row}:D{row}")
        row += 1

    # Freeze header row
    ws.freeze_panes = f"B{data_start}"

    # Print setup
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True


# ── Build all switch sheets ──────────────────────────────────────────────────
for sw in SWITCHES:
    build_switch_sheet(wb, sw)


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET: RESUMEN
# ═══════════════════════════════════════════════════════════════════════════
ws_res = wb.create_sheet("Resumen", 1)  # Insert at position 1 (after Instrucciones)
ws_res.sheet_view.showGridLines = False

col_ws_res = [2, 26, 18, 13, 13, 13, 13, 13, 20]
for i, w in enumerate(col_ws_res, 1):
    ws_res.column_dimensions[get_column_letter(i)].width = w

# Banner
for r in range(1, 6):
    ws_res.row_dimensions[r].height = 13
    for c in range(1, 12):
        ws_res.cell(row=r, column=c).fill = fill("1A1F2E")

ws_res.merge_cells("B2:I3")
t = ws_res["B2"]
t.value = "Resumen General — JSJ 2026 Switch Configuration"
t.font = ft(bold=True, size=16, color="FFFFFF")
t.alignment = Alignment(horizontal="left", vertical="center")

ws_res.merge_cells("B4:I4")
sub = ws_res["B4"]
sub.value = "Estadísticas consolidadas de los 3 switches · IBC Venues · 2026-04-07"
sub.font = ft(size=10, color="8899BB")
sub.alignment = Alignment(horizontal="left", vertical="center")

row = 6
ws_res.row_dimensions[row].height = 12
row += 1

# ── Summary Table per switch ──
ws_res.row_dimensions[row].height = 22
sum_headers = ["Switch", "Modelo", "IP Gestión", "Activos", "Vacíos", "Desasig.", "Reserv.", "Total", "Ubicación"]
for ci, h in enumerate(sum_headers, 2):
    c = ws_res.cell(row=row, column=ci)
    c.value = h
    c.font = ft(bold=True, size=10, color="FFFFFF")
    c.fill = fill("1A1F2E")
    c.alignment = center()
    c.border = border_thin()
row += 1

# Data per switch — link to each sheet
for si, sw in enumerate(SWITCHES):
    ws_res.row_dimensions[row].height = 20
    bg = "F7F8FA" if si % 2 == 0 else "FFFFFF"
    n = len(sw["puertos"])
    activos    = sum(1 for p in sw["puertos"] if p[3]=="activo")
    vacios     = sum(1 for p in sw["puertos"] if p[3]=="vacio")
    desasig    = sum(1 for p in sw["puertos"] if p[3]=="desasignado")
    reserv     = sum(1 for p in sw["puertos"] if p[3]=="reservado")
    vals = [sw["nombre"], sw["modelo"], sw["ip"], activos, vacios, desasig, reserv, n, sw["ubicacion"]]
    for ci, val in enumerate(vals, 2):
        c = ws_res.cell(row=row, column=ci)
        c.value = val
        c.border = border_thin()
        c.alignment = center() if ci > 4 else left_mid()
        c.fill = fill(bg)
        c.font = ft(bold=(ci==2), size=10)
    row += 1

# Totals row
ws_res.row_dimensions[row].height = 22
tot_lbl = ws_res.cell(row=row, column=2, value="TOTALES")
tot_lbl.font = ft(bold=True, size=10, color="FFFFFF")
tot_lbl.fill = fill("185FA5")
tot_lbl.border = border_thin()
tot_lbl.alignment = center()
ws_res.merge_cells(f"B{row}:D{row}")

total_ports = sum(len(sw["puertos"]) for sw in SWITCHES)
total_activos = sum(sum(1 for p in sw["puertos"] if p[3]=="activo") for sw in SWITCHES)
total_vacios  = sum(sum(1 for p in sw["puertos"] if p[3]=="vacio") for sw in SWITCHES)
total_desasig = sum(sum(1 for p in sw["puertos"] if p[3]=="desasignado") for sw in SWITCHES)
total_reserv  = sum(sum(1 for p in sw["puertos"] if p[3]=="reservado") for sw in SWITCHES)
totals = [total_activos, total_vacios, total_desasig, total_reserv, total_ports]

for ci, val in enumerate(totals, 5):
    c = ws_res.cell(row=row, column=ci)
    c.value = val
    c.font = ft(bold=True, size=10, color="FFFFFF")
    c.fill = fill("185FA5")
    c.alignment = center()
    c.border = border_thin()
row += 1

# ── VLAN Distribution across all switches ──
row += 1
ws_res.row_dimensions[row].height = 12
row += 1

ws_res.merge_cells(f"B{row}:I{row}")
vh = ws_res.cell(row=row, column=2)
vh.value = "Distribución por VLAN (todos los switches)"
vh.font = ft(bold=True, size=11, color="FFFFFF")
vh.fill = fill("1A1F2E")
vh.alignment = center()
row += 1

ws_res.row_dimensions[row].height = 20
vlan_hdrs = ["VLAN", "Descripción", "FSW 124F", "624F IBC A", "624F IBC B", "Total", "Prioridad", "Color"]
for ci, h in enumerate(vlan_hdrs, 2):
    c = ws_res.cell(row=row, column=ci)
    c.value = h
    c.font = ft(bold=True, size=10, color="FFFFFF")
    c.fill = fill("2D3250")
    c.alignment = center()
    c.border = border_thin()
row += 1

for vi, (vname, vinfo) in enumerate(VLAN_MAP.items()):
    bg = "F7F8FA" if vi % 2 == 0 else "FFFFFF"
    ws_res.row_dimensions[row].height = 20
    counts = [count_vlan(sw, vname) for sw in SWITCHES]
    total  = sum(counts)
    prio   = PRIORITY.get(vname, "—")

    vals = [vname, vinfo["desc"]] + counts + [total, prio]
    for ci, val in enumerate(vals, 2):
        c = ws_res.cell(row=row, column=ci)
        c.value = val
        c.border = border_thin()
        if ci == 2:  # VLAN name
            c.font = ft(bold=True, size=10, color=vinfo["color"])
            c.fill = fill(bg)
            c.alignment = left_mid()
        elif ci == 3:  # Desc
            c.font = ft(size=10)
            c.fill = fill(bg)
            c.alignment = left_mid()
        elif ci == 9:  # Color swatch
            pass  # handled below
        else:
            c.font = ft(size=10)
            c.fill = fill(bg)
            c.alignment = center()

    # Color swatch in last column
    sc = ws_res.cell(row=row, column=10)
    sc.value = vinfo["short"]
    sc.font = ft(bold=True, size=9, color="FFFFFF")
    sc.fill = fill(vinfo["color"])
    sc.alignment = center()
    sc.border = border_thin()
    row += 1


# ── Save ─────────────────────────────────────────────────────────────────────
out_path = "/Users/josuellorente/web/JSJ2026_Switches_Config.xlsx"
wb.save(out_path)
print(f"✅ Saved: {out_path}")
